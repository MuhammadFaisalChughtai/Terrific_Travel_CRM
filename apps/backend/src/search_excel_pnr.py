import openpyxl

wb = openpyxl.load_workbook('e:/Terrific-Travel_TMS/Polani_Sheet.xlsx')
sheet = wb.active

pnrs = ['HFM4WQ', 'HF7B5M', 'HF75SV', 'OMEOJY', 'H8D50S', 'FV63LM']
print('Searching Excel for PNRs:', pnrs)

# Print header row
header = [sheet.cell(1, c).value for c in range(1, 12)]
print('Header:', header)

found_rows = []
for r in range(2, sheet.max_row + 1):
    pnr_val = sheet.cell(r, 2).value
    if pnr_val and any(p in str(pnr_val).upper() for p in pnrs):
        row_vals = [sheet.cell(r, c).value for c in range(1, 12)]
        found_rows.append((r, row_vals))

print(f'Found {len(found_rows)} row(s):')
for r_num, vals in found_rows:
    print(f'Row {r_num}: {vals}')
